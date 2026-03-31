import openpyxl
from datetime import datetime

def amount_to_words(amount_str):
    amount = int(str(amount_str).replace(",", "").replace("Rs.", "").strip())
    
    ones = ["", "one", "two", "three", "four", "five", "six", "seven", "eight", "nine",
            "ten", "eleven", "twelve", "thirteen", "fourteen", "fifteen", "sixteen",
            "seventeen", "eighteen", "nineteen"]
    tens = ["", "", "twenty", "thirty", "forty", "fifty", "sixty", "seventy", "eighty", "ninety"]

    def below_thousand(n):
        if n == 0:
            return ""
        elif n < 20:
            return ones[n]
        elif n < 100:
            return tens[n // 10] + ("" if n % 10 == 0 else " " + ones[n % 10])
        else:
            return ones[n // 100] + " hundred" + ("" if n % 100 == 0 else " " + below_thousand(n % 100))

    if amount == 0:
        return "zero rupees"

    result = ""
    if amount >= 100000:
        result += below_thousand(amount // 100000) + " lakh "
        amount %= 100000
    if amount >= 1000:
        result += below_thousand(amount // 1000) + " thousand "
        amount %= 1000
    if amount > 0:
        result += below_thousand(amount)

    return result.strip() + " rupees"

def format_date(date_str):
    date_str = str(date_str).strip()
    formats = ["%d/%m/%Y", "%m/%d/%Y", "%Y-%m-%d", "%d-%m-%Y", "%B %d %Y", "%b %d %Y"]
    
    for fmt in formats:
        try:
            dt = datetime.strptime(date_str, fmt)
            day = dt.day
            if 11 <= day <= 13:
                suffix = "th"
            elif day % 10 == 1:
                suffix = "st"
            elif day % 10 == 2:
                suffix = "nd"
            elif day % 10 == 3:
                suffix = "rd"
            else:
                suffix = "th"
            return dt.strftime(f"%B {day}{suffix}, %Y")
        except:
            continue
    return date_str

def load_customers(filepath="customers.xlsx"):
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    customers = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        customer = dict(zip(headers, row))

        # Format for prompt
        customer["amount_spoken"] = amount_to_words(customer["amount"])
        customer["due_date_spoken"] = format_date(customer["due_date"])
        customer["customer_name"] = str(customer["customer_name"]).title()
        customer["attempts"] = str(customer.get("attempts") or "1")
        customer["payment_plan_eligible"] = str(customer.get("payment_plan_eligible") or "yes").lower()
        customer["notes"] = str(customer.get("notes") or "none")

        customers.append(customer)

    return customers

if __name__ == "__main__":
    customers = load_customers()
    for c in customers:
        print(c)